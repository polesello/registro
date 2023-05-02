from django.shortcuts import render, redirect
from django.http import HttpResponse

from .models import *


def index(request):
    from random import randint
    totale_gatti = randint(0, 44)
    lista_gatti = range(totale_gatti)
    persone = Persona.objects.all()

    context = {
        'nome': 'Nello',
        'cognome': 'Polesello',
        'totale_gatti': totale_gatti,
        'lista_gatti': lista_gatti,
        'presenti': 6,
        'persone': persone,
    }

    print(context)
    return render(request, 'homepage.html', context)


def index_old(request):
    if 'one' in request.path:
        return HttpResponse("<h1 style='color:red'>Ciaone</h1>")
    else:
        return HttpResponse("<h4>Ciao</h4>")
    


def mostra_targa(request, targa):

    prossima = 'cc123aa'
    valid = targa[:2].isalpha() and targa[2:5].isdigit() and targa[5:].isalpha()
    square = False
    if len(targa) == 7 and valid:
        targa = targa.upper()

        prossima = targa[:2] + str(int(targa[2:5]) + 1).zfill(3) + targa[5:]

        targa = targa[:2] + ' &nbsp; ' + targa[2:]
        square = targa.startswith('Z')

    context = {
        'targa': targa,
        'valid': valid,
        'square': square,
        'prossima': prossima,
    }
    return render(request, 'targa.html', context)


def corsi(request):
    context = {
        'corsi': Corso.objects.all(),
    }
    return render(request, 'corsi.html', context)


def lezioni(request, id):
    import openpyxl

    corso = Corso.objects.get(id=id)
    lezioni = Lezione.objects.filter(materia__modulo__corso=corso).order_by('data')

    if 'excel' in request.GET:

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = corso.titolo

        c = ws.cell(row=2, column=3)
        c.value = 'CIAO'

        for i, lezione in enumerate(lezioni):
            c = ws.cell(row=i+1, column=1) 
            c.value = lezione.argomento

            c = ws.cell(row=i+1, column=2) 
            c.value = lezione.materia.nome

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'inline; filename=calendario.xlsx'
        
        wb.save(response)

        return response
    
    context = {
        'corso': corso,
        'lezioni': lezioni
    }
    return render(request, 'lezioni.html', context)


def dettaglio_lezione(request, id):
    from .forms import LezioneForm

    lezione = Lezione.objects.get(id=id)

    form = LezioneForm(request.POST or None, instance=lezione)

    presenti = request.POST.getlist('presente')
    print('presenti ' , presenti )

    if request.method == 'POST':

        # Salva i dati della lezione
        if 'salva-dettagli' in request.POST:
            if form.is_valid():
                form.save()
            





        # Salva i presenti alla lezione
        if 'salva-presenze' in request.POST:
            Presenza.objects.filter(lezione=lezione).delete()

            for id_persona in presenti:
                presenza = Presenza(lezione=lezione, persona_id=id_persona)
                presenza.save()

            return redirect(request.path)
        



    persone_presenti = Persona.objects.filter(presenze__lezione = lezione)
    context = {'lezione': lezione, 'persone_presenti':persone_presenti, 'form':form}
    return render(request, 'dettaglio_lezione.html', context)
